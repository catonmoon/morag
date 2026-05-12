"""Конвертер xlsx ↔ json для запуска eval_native.py внутри контейнера.

Использование:
    python eval_convert.py to-json   <input.xlsx> <output.json>
    python eval_convert.py from-json <input.json> <output.xlsx>
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl


def to_json(xlsx_path: Path, json_path: Path) -> None:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    rows = []
    for r in range(2, ws.max_row + 1):
        q = ws.cell(r, 3).value
        if not q:
            continue
        rows.append({
            'num': ws.cell(r, 1).value,
            'section': ws.cell(r, 2).value,
            'question': q,
            'gold': ws.cell(r, 4).value,
        })
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(rows, f, ensure_ascii=False, indent=1)
    print(f'{xlsx_path} → {json_path}: {len(rows)} questions')


def from_json(json_path: Path, xlsx_path: Path) -> None:
    with open(json_path, 'r', encoding='utf-8') as f:
        rows = json.load(f)
    rows.sort(key=lambda x: (x.get('num') or 0))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Results'
    headers = [
        'Номер', 'Раздел', 'Вопрос', 'Правильный ответ',
        'Ответ модели', 'Время ответа (с)', 'Документов найдено',
        'Tool-calls', 'Итераций', 'Ошибка', 'Timestamp',
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h)
    for i, r in enumerate(rows, 2):
        ws.cell(i, 1, r.get('num'))
        ws.cell(i, 2, r.get('section'))
        ws.cell(i, 3, r.get('question'))
        ws.cell(i, 4, r.get('gold'))
        ws.cell(i, 5, r.get('answer'))
        ws.cell(i, 6, r.get('elapsed_sec'))
        ws.cell(i, 7, r.get('docs'))
        ws.cell(i, 8, r.get('tool_calls'))
        ws.cell(i, 9, r.get('iters'))
        ws.cell(i, 10, r.get('error'))
        ws.cell(i, 11, r.get('timestamp'))
    wb.save(xlsx_path)
    print(f'{json_path} → {xlsx_path}: {len(rows)} rows')


if __name__ == '__main__':
    if len(sys.argv) < 4 or sys.argv[1] not in ('to-json', 'from-json'):
        print(__doc__)
        sys.exit(2)
    mode, in_path, out_path = sys.argv[1], Path(sys.argv[2]), Path(sys.argv[3])
    if mode == 'to-json':
        to_json(in_path, out_path)
    else:
        from_json(in_path, out_path)
